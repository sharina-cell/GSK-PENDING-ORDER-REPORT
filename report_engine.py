"""
GSK Pending Order Report — Core Engine
Generates a formatted Excel workbook from TC, Shopee, WMS, and Inventory files.
"""

import io
import warnings
import zipfile
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Styles ────────────────────────────────────────────────────────────────────
_thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HDR_FILL    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", size=10, bold=True)

NICK_FILLS = {
    "shopee": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "tiktok": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "lazada": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
}
NICK_FILLS_ALT = {
    "shopee": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "tiktok": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "lazada": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
}
WMS_FILLS = {
    "FULFILLED":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PENDING_FULFILLMENT": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "PARTIAL_ALLOCATED":   PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NONE_ALLOCATED":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CANCEL":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "RETURN":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CLOSE":               PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "NOT FOUND":           PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
RED_C       = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YEL_C       = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLU_C       = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ORANGE_ALT  = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

COL_LABELS = [
    "Order Number", "Invoice Number", "Payment Status",
    "Order Item Status", "Ordered Date", "Nickname", "MP SLA", "WMS",
]

# ── ZIP Handling ──────────────────────────────────────────────────────────────
def _load_wms(wms_file) -> pd.DataFrame:
    """
    Load the WMS file. Accepts either a single .xlsx file or a .zip
    containing multiple SalesOrder_Item_Report .xlsx parts — these are
    concatenated automatically.
    """
    name = getattr(wms_file, "name", str(wms_file))
    if str(name).lower().endswith(".zip"):
        with zipfile.ZipFile(wms_file) as z:
            xlsx_names = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
            if not xlsx_names:
                raise ValueError("WMS zip contains no .xlsx files")
            frames = []
            for n in xlsx_names:
                with z.open(n) as f:
                    frames.append(pd.read_excel(io.BytesIO(f.read()), dtype=str))
            return pd.concat(frames, ignore_index=True)
    return pd.read_excel(wms_file, dtype=str)


# ── Date Helpers ──────────────────────────────────────────────────────────────
def _add_one_day(date_str: str) -> str:
    if not date_str or str(date_str).strip() in ("", "nan", "None"):
        return ""
    s = str(date_str).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d/%Y",
    ):
        try:
            return (datetime.strptime(s, fmt) + timedelta(days=1)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    for trim, fmt in [
        (19, "%d/%m/%Y %H:%M:%S"), (10, "%d/%m/%Y"),
        (19, "%Y-%m-%d %H:%M:%S"), (10, "%Y-%m-%d"),
    ]:
        try:
            return (datetime.strptime(s[:trim], fmt) + timedelta(days=1)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _fmt_short(d: str) -> str:
    if not d or str(d).strip() in ("", "nan"):
        return "No Date"
    s = str(d).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(s, fmt).date().strftime("%d-%b-%y")
        except ValueError:
            pass
    for trim, fmt in [
        (19, "%Y-%m-%d %H:%M:%S"), (16, "%Y-%m-%d %H:%M"), (10, "%Y-%m-%d"),
        (19, "%d/%m/%Y %H:%M:%S"), (10, "%d/%m/%Y"),
    ]:
        try:
            return datetime.strptime(s[:trim], fmt).date().strftime("%d-%b-%y")
        except ValueError:
            pass
    return s[:10]


def _parse_short(s: str):
    if s == "No Date":
        return datetime.max.date()
    try:
        return datetime.strptime(s, "%d-%b-%y").date()
    except ValueError:
        return datetime.max.date()


# ── Excel Helpers ─────────────────────────────────────────────────────────────
def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def _write_simple_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    _style_header(ws)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Main Engine ───────────────────────────────────────────────────────────────
def generate_report(
    tc_file,
    shopee_file,
    wms_file,
    inv_file=None,
    progress_callback=None,
) -> tuple:
    """
    Build the GSK Pending Order Report Excel workbook.

    Parameters
    ----------
    tc_file      : file-like object or path — GSK_TC_ORDER_REPORT.csv
    shopee_file  : file-like object or path — GSK_SHOPEE_ORDER.xlsx
    wms_file     : file-like object or path — GSK_WMS.xlsx OR GSK_WMS.zip
                   (zip files containing multiple SalesOrder_Item_Report parts
                   are automatically extracted and concatenated)
    inv_file     : file-like object or path — InventoryWarehouse_Report_*.xlsx (optional)
    progress_callback : callable(step: str) — optional UI progress hook

    Returns
    -------
    (bytes, dict) : the Excel workbook as bytes, and a summary dict with counts
    """

    def _progress(msg: str):
        if progress_callback:
            progress_callback(msg)

    today = datetime.today().date()

    # ── Load TC ───────────────────────────────────────────────────────────────
    _progress("Loading TC Order Report…")
    tc = pd.read_csv(tc_file, dtype=str)
    for col in tc.columns:
        tc[col] = tc[col].astype(str).str.replace("\r", "").str.strip()
    tc.replace("nan", "", inplace=True)

    # ── Load Shopee ───────────────────────────────────────────────────────────
    _progress("Loading Shopee orders…")
    shopee = pd.read_excel(shopee_file, dtype=str)
    shopee["Order ID"] = shopee["Order ID"].astype(str).str.strip()
    shopee["Estimated Ship Out Date"] = shopee["Estimated Ship Out Date"].astype(str).str.strip()
    shopee_map = (
        shopee.drop_duplicates("Order ID")
        .set_index("Order ID")["Estimated Ship Out Date"]
        .to_dict()
    )

    # ── Load WMS (handles zip automatically) ─────────────────────────────────
    _progress("Loading WMS…")
    wms = _load_wms(wms_file)
    wms["client_order_id"] = wms["client_order_id"].astype(str).str.strip()
    wms["_match_id"] = wms["client_order_id"].str.replace(r"^GSK_", "", regex=True)
    wms["status_so"] = wms["status_so"].astype(str).str.strip()
    wms_dedup = wms.drop_duplicates("_match_id")
    wms_map      = wms_dedup.set_index("_match_id")["status_so"].to_dict()
    wms_map_full = wms_dedup.set_index("client_order_id")["status_so"].to_dict()

    # ── Load Inventory ────────────────────────────────────────────────────────
    inv_map = {}
    inv_format = "none"
    if inv_file is not None:
        _progress("Loading Inventory…")
        inv = pd.read_excel(inv_file, dtype=str)
        if "client_code_item" in inv.columns:
            inv["quantity_available"] = inv["quantity_available"].apply(
                lambda x: "0" if str(x).strip() in ("nan", "", "None") else str(x).strip()
            )
            inv_map    = inv.set_index("client_code_item")["quantity_available"].to_dict()
            inv_format = "warehouse"
        elif "code_item" in inv.columns:
            inv["quantity_available"] = pd.to_numeric(inv["quantity_available"], errors="coerce").fillna(0)
            inv_map    = inv.groupby("code_item")["quantity_available"].sum().to_dict()
            inv_format = "zone"

    def _get_stock(raw_sku: str) -> int:
        if not raw_sku:
            return -1
        base = raw_sku.split("x")[0].strip()
        val  = (
            inv_map.get(base, inv_map.get(f"GSK_{base}", None))
            if inv_format == "warehouse"
            else inv_map.get(f"GSK_{base}", inv_map.get(base, None))
        )
        if val is None:
            return -1
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return -1

    def _get_inv_stock(sku: str) -> int:
        if not sku or str(sku).strip() in ("", "nan"):
            return -1
        val = inv_map.get(str(sku).strip())
        if val is None:
            return -1
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return -1

    # ── Filter ────────────────────────────────────────────────────────────────
    _progress("Filtering orders…")
    j_values      = ["ACCEPTED/PICKED", "NEW", "READY TO SHIP"]
    mask_j        = tc["order_item_status"].isin(j_values)
    mask_bc       = tc["nickname"].str.lower().str.contains("tiktok|lazada|shopee", na=False)
    mask_comp     = tc["payment_status"] == "COMPLETED"
    mask_pend_cod = (tc["payment_status"] == "PENDING") & tc["payment_methods"].str.contains("COD", na=False)
    filtered = tc[mask_j & (mask_comp | mask_pend_cod) & mask_bc].copy()

    # ── MP SLA ────────────────────────────────────────────────────────────────
    def _mp_sla(row) -> str:
        nick = str(row["nickname"]).lower()
        if "shopee" in nick:
            val = shopee_map.get(str(row["order_id"]).strip()) or \
                  shopee_map.get(str(row["order_number"]).strip()) or ""
            return str(val)[:10] if val and val not in ("nan", "None", "") else ""
        elif "tiktok" in nick or "lazada" in nick:
            return _add_one_day(row["ordered_date"])
        return ""

    filtered["MP SLA"] = filtered.apply(_mp_sla, axis=1)

    # ── WMS Lookup ────────────────────────────────────────────────────────────
    def _lookup_wms(inv_num: str) -> str:
        inv_num = str(inv_num).strip()
        r = wms_map.get(inv_num) or wms_map_full.get(inv_num) or wms_map_full.get(f"GSK_{inv_num}")
        return r if r and r not in ("nan", "None", "") else "NOT FOUND"

    filtered["WMS"]    = filtered["invoice_number"].apply(_lookup_wms)
    partial_none_mask  = filtered["WMS"].isin(["PARTIAL_ALLOCATED", "NONE_ALLOCATED"])
    filtered["_stock"] = filtered["custom_sku"].apply(_get_stock)
    zero_stock_partial = partial_none_mask & (filtered["_stock"] == 0)

    # ── Sheet DataFrames ──────────────────────────────────────────────────────
    COLS   = ["order_number", "invoice_number", "payment_status",
              "order_item_status", "ordered_date", "nickname", "MP SLA", "WMS"]
    result = filtered[COLS].copy()
    result.columns = COL_LABELS

    # FILTERED DATA: exclude NOT FOUND, zero-stock partial, and CANCEL/CANCELLED
    cancel_mask = filtered["WMS"].str.upper().isin(["CANCEL", "CANCELLED"])
    excl_mask   = (filtered["WMS"] == "NOT FOUND") | zero_stock_partial | cancel_mask
    fd_df  = result[~excl_mask].copy()
    np_df  = result[filtered["WMS"] == "NOT FOUND"].copy()
    cnf_df = result[filtered["WMS"].isin(["CLOSE", "NOT FOUND"])].copy()

    # ── PARTIAL & NON ALLOCATED ───────────────────────────────────────────────
    # TC has one row per SKU for multi-item orders — dedupe to ONE row per
    # invoice for the order-summary columns; WMS line items are expanded
    # separately so each SKU still gets its own row in the output sheet.
    partial_rows = filtered[partial_none_mask].copy()
    partial_rows_dedup = partial_rows.drop_duplicates(subset=["invoice_number"]).copy()

    partial_df = partial_rows_dedup[COLS].copy()
    partial_df.columns = COL_LABELS
    partial_df["Custom SKU"] = partial_rows_dedup["custom_sku"].apply(
        lambda r: f"GSK_{r.split('x')[0].strip()}" if r else ""
    )
    if inv_file is not None:
        partial_df["WH Stock"] = partial_rows_dedup["_stock"].apply(lambda x: "" if x == -1 else x)
        partial_df["Remarks"]  = partial_rows_dedup["_stock"].apply(
            lambda x: "CANCEL" if x == 0 else ("FULFILL" if x > 0 else "")
        )
    else:
        partial_df["WH Stock"] = ""
        partial_df["Remarks"]  = ""

    # WMS line items per partial order, with per-SKU remarks logic:
    #   1. Single SKU, zero allocated total            → CANCEL
    #   2. Multiple SKUs, this SKU zero allocated       → REQUEST REFUND
    #      Multiple SKUs, this SKU fully allocated      → FULFILL
    #   3. Same SKU split across lines, under-allocated → REQUEST PARTIAL REFUND
    partial_inv_nums  = partial_rows["invoice_number"].unique().tolist()
    wms_partial_lines = wms[wms["_match_id"].isin(partial_inv_nums)].copy()
    partial_sku_detail = {}
    for inv_num in partial_inv_nums:
        inv_num_clean = str(inv_num).strip()
        lines = wms_partial_lines[wms_partial_lines["_match_id"] == inv_num_clean]
        items = []
        for _, line in lines.iterrows():
            sku           = str(line.get("client_code_item", "")).strip()
            descr         = str(line.get("item_descr", "")).strip()
            qty_ord_raw   = str(line.get("quantity", "")).strip()
            qty_alloc_raw = str(line.get("quantity_allocated", "")).strip()
            if qty_alloc_raw in ("nan", "None", ""):
                qty_alloc_raw = "0"
            wh_stock = _get_inv_stock(sku) if inv_file is not None else -1
            try:
                qty_ord_int = int(float(qty_ord_raw))
            except (ValueError, TypeError):
                qty_ord_int = 0
            try:
                qty_alloc_int = int(float(qty_alloc_raw))
            except (ValueError, TypeError):
                qty_alloc_int = 0
            items.append({
                "SKU": sku, "Description": descr,
                "Qty Ordered": qty_ord_raw, "Qty Allocated": qty_alloc_raw,
                "Qty Ordered Int": qty_ord_int, "Qty Allocated Int": qty_alloc_int,
                "WH Stock": "" if wh_stock == -1 else wh_stock,
                "Has Stock": wh_stock > 0 if wh_stock != -1 else None,
                "Remarks": "",
            })

        num_skus    = len(set(i["SKU"] for i in items))
        total_ord   = sum(i["Qty Ordered Int"]   for i in items)
        total_alloc = sum(i["Qty Allocated Int"] for i in items)

        for item in items:
            qty_ord   = item["Qty Ordered Int"]
            qty_alloc = item["Qty Allocated Int"]
            is_zero   = qty_alloc == 0

            if num_skus == 1:
                if total_alloc == 0:
                    item["Remarks"] = "CANCEL"
                elif total_alloc < total_ord:
                    item["Remarks"] = "REQUEST PARTIAL REFUND"
                else:
                    item["Remarks"] = "FULFILL"
            else:
                if is_zero:
                    item["Remarks"] = "REQUEST REFUND"
                elif qty_alloc >= qty_ord:
                    item["Remarks"] = "FULFILL"
                else:
                    item["Remarks"] = "REQUEST PARTIAL REFUND"

        partial_sku_detail[inv_num_clean] = items

    # RMA
    rma_tc = tc[tc["order_status"].isin(["CANCELLED", "CANCEL REQUESTED"])].copy()
    rma_tc["WMS"] = rma_tc["invoice_number"].apply(_lookup_wms)
    rma_tc = rma_tc[rma_tc["WMS"] == "FULFILLED"].drop_duplicates("invoice_number")
    rma_df = rma_tc[["order_number", "invoice_number", "payment_status",
                      "order_item_status", "ordered_date", "nickname", "WMS"]].copy()
    rma_df.columns = ["Order Number", "Invoice Number", "Payment Status",
                      "Order Item Status", "Ordered Date", "Nickname", "WMS"]

    # ── Pivot ─────────────────────────────────────────────────────────────────
    _progress("Building pivot table…")
    fd_df["_SLA_short"] = fd_df["MP SLA"].apply(_fmt_short)
    pivot = fd_df.groupby(["Order Item Status", "_SLA_short"]).size().unstack(fill_value=0)
    pivot = pivot[sorted(pivot.columns, key=_parse_short)]
    pivot["Grand Total"] = pivot.sum(axis=1)
    grand_row = pivot.sum(axis=0)
    grand_row.name = "Grand Total"
    pivot = pd.concat([pivot, grand_row.to_frame().T])

    # ── Write Excel ───────────────────────────────────────────────────────────
    _progress("Writing Excel workbook…")
    wb = Workbook()
    wb.remove(wb.active)

    # FILTERED DATA
    ws_fd = wb.create_sheet("FILTERED DATA")
    ws_fd.sheet_properties.tabColor = "4472C4"
    ws_fd.append(COL_LABELS)
    _style_header(ws_fd)
    nick_counter: dict = {}
    for _, row in fd_df.iterrows():
        nick = str(row["Nickname"]).lower()
        mp   = "shopee" if "shopee" in nick else ("tiktok" if "tiktok" in nick else ("lazada" if "lazada" in nick else ""))
        nick_counter[mp] = nick_counter.get(mp, 0) + 1
        row_fill = (NICK_FILLS_ALT.get(mp) if nick_counter[mp] % 2 == 0 else NICK_FILLS.get(mp)) if mp else None
        ws_fd.append(list(row[COL_LABELS]))
        xl_row = ws_fd.max_row
        for i, cell in enumerate(ws_fd[xl_row]):
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if COL_LABELS[i] == "WMS":
                wf = WMS_FILLS.get(str(cell.value))
                if wf:
                    cell.fill = wf
            elif row_fill:
                cell.fill = row_fill
    ws_fd.freeze_panes = "A2"
    _auto_width(ws_fd)

    # PIVOT
    ws_pv = wb.create_sheet("PIVOT")
    ws_pv.sheet_properties.tabColor = "4472C4"
    pv_cols = ["Order Item Status"] + list(pivot.columns)
    ws_pv.append(pv_cols)
    for i, cell in enumerate(ws_pv[1]):
        cell.font      = HDR_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        h = pv_cols[i]
        if i == 0 or h == "Grand Total":
            cell.fill = HDR_FILL
        else:
            try:
                col_date = datetime.strptime(h, "%d-%b-%y").date()
                if col_date < today:
                    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                elif col_date == today:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(bold=True, color="000000", name="Arial", size=10)
                else:
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            except ValueError:
                cell.fill = HDR_FILL

    for status, prow in pivot.iterrows():
        data_row = [status] + [prow.get(c, 0) for c in pivot.columns]
        ws_pv.append(data_row)
        xl_row   = ws_pv.max_row
        is_grand = status == "Grand Total"
        for i, cell in enumerate(ws_pv[xl_row]):
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if i > 0 else "left", vertical="center")
            if is_grand:
                cell.font = BOLD_FONT
                cell.fill = GREY_FILL
            else:
                cell.font = BODY_FONT
                if i > 0:
                    h = pv_cols[i]
                    if h == "Grand Total":
                        cell.fill = GREY_FILL
                    else:
                        try:
                            col_date  = datetime.strptime(h, "%d-%b-%y").date()
                            cell.fill = RED_C if col_date < today else (YEL_C if col_date == today else BLU_C)
                        except ValueError:
                            pass
    ws_pv.freeze_panes = "A2"
    _auto_width(ws_pv)

    # NOT PUSHED
    ws_np = wb.create_sheet("NOT PUSHED")
    ws_np.sheet_properties.tabColor = "C00000"
    _write_simple_sheet(ws_np, np_df)

    # PARTIAL & NON ALLOCATED
    ws_pa = wb.create_sheet("PARTIAL & NON ALLOCATED")
    ws_pa.sheet_properties.tabColor = "ED7D31"
    pa_labels = COL_LABELS + ["WMS Custom SKU", "Item Description", "Qty Ordered", "Qty Allocated", "WH Stock", "Remarks"]
    ws_pa.append(pa_labels)
    _style_header(ws_pa)
    order_toggle: dict = {}
    for _, row in partial_df.iterrows():
        inv_num = str(row["Invoice Number"]).strip()
        items   = partial_sku_detail.get(inv_num, [])
        if inv_num not in order_toggle:
            order_toggle[inv_num] = len(order_toggle) % 2 == 0
        base_fill = ORANGE_ALT if order_toggle[inv_num] else ORANGE_FILL

        if not items:
            data_row = list(row[COL_LABELS]) + [row["Custom SKU"], "", "", "", row["WH Stock"], row["Remarks"]]
            ws_pa.append(data_row)
            xl_row = ws_pa.max_row
            remark = str(row["Remarks"])
            for i, cell in enumerate(ws_pa[xl_row]):
                cell.font = BODY_FONT; cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                col_name = pa_labels[i]
                if col_name == "Remarks":
                    cell.fill = RED_FILL if remark == "CANCEL" else (GREEN_FILL if remark == "FULFILL" else base_fill)
                elif col_name == "WMS":
                    cell.fill = WMS_FILLS.get(str(cell.value), base_fill)
                else:
                    cell.fill = base_fill
        else:
            for item_idx, item in enumerate(items):
                order_cols = list(row[COL_LABELS]) if item_idx == 0 else [""] * len(COL_LABELS)
                wh_stock   = item["WH Stock"]
                has_stock  = item["Has Stock"]
                remark     = item.get("Remarks", "")
                data_row   = order_cols + [item["SKU"], item["Description"], item["Qty Ordered"], item["Qty Allocated"], wh_stock, remark]
                ws_pa.append(data_row)
                xl_row = ws_pa.max_row
                for i, cell in enumerate(ws_pa[xl_row]):
                    cell.font = BODY_FONT; cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                    col_name = pa_labels[i]
                    if col_name == "Remarks":
                        if remark == "CANCEL":
                            cell.fill = RED_FILL
                        elif remark == "FULFILL":
                            cell.fill = GREEN_FILL
                        elif remark == "REQUEST REFUND":
                            cell.fill = RED_FILL
                        elif remark == "REQUEST PARTIAL REFUND":
                            cell.fill = YELLOW_FILL
                        else:
                            cell.fill = base_fill
                    elif col_name == "WH Stock" and inv_file is not None:
                        cell.fill = GREEN_FILL if has_stock is True else (RED_FILL if has_stock is False else base_fill)
                    elif col_name == "WMS" and item_idx == 0:
                        cell.fill = WMS_FILLS.get(str(cell.value), base_fill)
                    else:
                        cell.fill = base_fill
                ws_pa.row_dimensions[xl_row].height = 16
    ws_pa.freeze_panes = "A2"
    _auto_width(ws_pa)

    # CLOSE & NOT FOUND
    ws_cnf = wb.create_sheet("CLOSE & NOT FOUND")
    ws_cnf.sheet_properties.tabColor = "808080"
    _write_simple_sheet(ws_cnf, cnf_df)

    # ORIGINAL
    ws_orig = wb.create_sheet("ORIGINAL (All Data)")
    ws_orig.sheet_properties.tabColor = "1F2D3D"
    ws_orig.append(list(tc.columns))
    _style_header(ws_orig)
    for row in tc.itertuples(index=False):
        ws_orig.append(list(row))
    for row in ws_orig.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
    ws_orig.freeze_panes = "A2"
    _auto_width(ws_orig)

    # RMA
    ws_rma = wb.create_sheet("RMA")
    ws_rma.sheet_properties.tabColor = "C00000"
    _write_simple_sheet(ws_rma, rma_df)

    # ── Return as bytes ───────────────────────────────────────────────────────
    _progress("Done!")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    summary = {
        "filtered":   len(fd_df),
        "not_pushed": len(np_df),
        "partial":    len(partial_df),
        "close_nf":   len(cnf_df),
        "rma":        len(rma_df),
    }
    return buffer.read(), summary
