import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GSK Pending Order Report",
    page_icon="📦",
    layout="wide",
)

# ─── STYLES ──────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
BODY_FONT = Font(name='Arial', size=10)
CENTER    = Alignment(horizontal='center', vertical='center')
LEFT      = Alignment(horizontal='left',   vertical='center')
thin      = Side(style='thin', color='BFBFBF')
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

NICK_FILLS = {
    'shopee': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    'tiktok': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    'lazada': PatternFill(start_color='EDE7F6', end_color='EDE7F6', fill_type='solid'),
}
NICK_FILLS_ALT = {
    'shopee': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
    'tiktok': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
    'lazada': PatternFill(start_color='D1C4E9', end_color='D1C4E9', fill_type='solid'),
}
WMS_FILLS = {
    'FULFILLED':           PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'PENDING_FULFILLMENT': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'PARTIAL_ALLOCATED':   PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'NONE_ALLOCATED':      PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'CANCEL':              PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'RETURN':              PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'CLOSE':               PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    'NOT FOUND':           PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
}
RED_FILL   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def nick_key(nick_str):
    n = str(nick_str or '').lower()
    for k in ['shopee', 'tiktok', 'lazada']:
        if k in n:
            return k
    return 'shopee'

def get_row_fill(nick_str, row_idx):
    k = nick_key(nick_str)
    return NICK_FILLS_ALT[k] if row_idx % 2 == 0 else NICK_FILLS[k]

def apply_row_ws(ws, row_num, values, row_fill, wms_col_idx=8, wms_val=None):
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=v)
        c.font   = BODY_FONT
        c.alignment = LEFT
        c.border = BORDER
        if wms_col_idx and ci == wms_col_idx and wms_val and wms_val in WMS_FILLS:
            c.fill = WMS_FILLS[wms_val]
        else:
            c.fill = row_fill

def write_header(ws, headers, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = CENTER
        c.border    = BORDER
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes      = 'A2'
    ws.row_dimensions[1].height = 20

def add_one_day(date_str):
    if not date_str:
        return ''
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
        try:
            d = datetime.strptime(date_str[:19].strip(), fmt)
            return (d + timedelta(days=1)).strftime('%Y-%m-%d')
        except:
            pass
    return ''

def get_stock(raw_sku, inv_map):
    if not raw_sku:
        return -1
    sku_base = raw_sku.split('x')[0]
    val = inv_map.get(f'GSK_{sku_base}', inv_map.get(sku_base, None))
    if val is None:
        return -1
    try:
        return int(float(val))
    except:
        return -1

def fmt_date(d_str):
    try:
        return datetime.strptime(d_str, '%Y-%m-%d').strftime('%d-%b-%y')
    except:
        return d_str

def build_pivot_sheet(ws, final_rows, headers, today_str):
    """Rebuild PIVOT sheet from final_rows list."""
    sla_idx    = headers.index('MP SLA')
    status_idx = headers.index('Order Item Status')

    from collections import defaultdict
    pivot = defaultdict(lambda: defaultdict(int))
    dates_set, statuses_set = set(), set()
    for rv in final_rows:
        sla = rv[sla_idx] or ''
        st  = rv[status_idx] or ''
        if sla:
            pivot[st][sla] += 1
            dates_set.add(sla)
            statuses_set.add(st)

    col_dates = sorted(dates_set)
    statuses  = sorted(statuses_set)

    PIVOT_RED  = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    PIVOT_YEL  = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    PIVOT_BLUE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    # Clear
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.value  = None
            cell.fill   = PatternFill()
            cell.border = Border()
            cell.font   = BODY_FONT

    ws.freeze_panes = 'B2'
    ws.cell(1, 1, 'Order Item Status').fill      = HDR_FILL
    ws.cell(1, 1).font      = HDR_FONT
    ws.cell(1, 1).alignment = CENTER
    ws.cell(1, 1).border    = BORDER
    ws.column_dimensions['A'].width = 24

    for ci, d in enumerate(col_dates, 2):
        c = ws.cell(1, ci, fmt_date(d))
        c.alignment = CENTER
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 14
        if d < today_str:
            c.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        elif d == today_str:
            c.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            c.font = Font(bold=True, color='000000', name='Arial', size=10)
        else:
            c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)

    gt_ci = len(col_dates) + 2
    ws.cell(1, gt_ci, 'Grand Total').fill      = HDR_FILL
    ws.cell(1, gt_ci).font      = HDR_FONT
    ws.cell(1, gt_ci).alignment = CENTER
    ws.cell(1, gt_ci).border    = BORDER
    ws.column_dimensions[get_column_letter(gt_ci)].width = 14

    grand_totals = [0] * len(col_dates)
    for ri, st in enumerate(statuses, 2):
        ws.cell(ri, 1, st).font  = BODY_FONT
        ws.cell(ri, 1).fill      = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        ws.cell(ri, 1).border    = BORDER
        row_total = 0
        for ci, d in enumerate(col_dates, 2):
            val = pivot[st][d]
            grand_totals[ci - 2] += val
            row_total += val
            c = ws.cell(ri, ci, val if val else '')
            c.alignment = CENTER
            c.font      = BODY_FONT
            c.border    = BORDER
            if d < today_str:
                c.fill = PIVOT_RED
            elif d == today_str:
                c.fill = PIVOT_YEL
            else:
                c.fill = PIVOT_BLUE
        gt = ws.cell(ri, gt_ci, row_total)
        gt.alignment = CENTER
        gt.font      = Font(bold=True, name='Arial', size=10)
        gt.border    = BORDER

    gt_row    = len(statuses) + 2
    final_grand = sum(grand_totals)
    ws.cell(gt_row, 1, 'Grand Total').fill   = HDR_FILL
    ws.cell(gt_row, 1).font   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    ws.cell(gt_row, 1).border = BORDER
    for ci, v in enumerate(grand_totals, 2):
        c = ws.cell(gt_row, ci, v if v else '')
        c.alignment = CENTER
        c.fill      = HDR_FILL
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.border    = BORDER
    ws.cell(gt_row, gt_ci, final_grand).fill      = HDR_FILL
    ws.cell(gt_row, gt_ci).font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    ws.cell(gt_row, gt_ci).alignment = CENTER
    ws.cell(gt_row, gt_ci).border    = BORDER

    return final_grand


# ─── CORE REPORT ENGINE ──────────────────────────────────────────────────────

def generate_report(tc_file, shopee_file, wms_files, inv_file):
    """Generate full GSK report. Returns (wb, summary_dict)."""

    # TC
    tc = pd.read_csv(tc_file, dtype=str)
    for col in tc.columns:
        tc[col] = tc[col].astype(str).str.replace('\r', '').str.strip()
    tc.replace('nan', '', inplace=True)

    # Shopee
    shopee = pd.read_excel(shopee_file, dtype=str)
    shopee['Order ID'] = shopee['Order ID'].astype(str).str.strip()
    shopee['Estimated Ship Out Date'] = shopee['Estimated Ship Out Date'].astype(str).str.strip()
    shopee_map = shopee.drop_duplicates('Order ID').set_index('Order ID')['Estimated Ship Out Date'].to_dict()

    # WMS (one or many)
    wms_dfs = []
    for f in wms_files:
        try:
            wms_dfs.append(pd.read_excel(f, dtype=str))
        except Exception as e:
            st.warning(f"Could not read WMS file: {e}")
    wms = pd.concat(wms_dfs, ignore_index=True) if wms_dfs else pd.DataFrame()
    wms['client_order_id'] = wms['client_order_id'].astype(str).str.strip()
    wms['status_so']       = wms['status_so'].astype(str).str.strip()
    wms_map = wms.drop_duplicates('client_order_id').set_index('client_order_id')['status_so'].to_dict()

    # Inventory
    inv = pd.read_excel(inv_file, dtype=str)
    inv_map = {}
    if 'client_code_item' in inv.columns:
        inv['quantity_available'] = inv['quantity_available'].apply(
            lambda x: '0' if str(x).strip() in ('nan', '', 'None') else str(x).strip()
        )
        inv_map = inv.set_index('client_code_item')['quantity_available'].to_dict()
    elif 'code_item' in inv.columns:
        inv['quantity_available'] = pd.to_numeric(inv['quantity_available'], errors='coerce').fillna(0)
        inv_map = inv.groupby('code_item')['quantity_available'].sum().to_dict()

    # ─ Filter ─
    j_values   = ['ACCEPTED/PICKED', 'NEW', 'READY TO SHIP']
    mask_j     = tc['order_item_status'].isin(j_values)
    mask_bc    = tc['nickname'].str.lower().str.contains('tiktok|lazada|shopee', na=False)
    mask_comp  = tc['payment_status'] == 'COMPLETED'
    mask_pcod  = (tc['payment_status'] == 'PENDING') & tc['payment_methods'].str.contains('COD', na=False)
    filtered   = tc[mask_j & (mask_comp | mask_pcod) & mask_bc].copy()

    def get_mp_sla(row):
        nick = str(row['nickname']).lower()
        if 'shopee' in nick:
            val = shopee_map.get(str(row['order_id']).strip()) or \
                  shopee_map.get(str(row['order_number']).strip()) or ''
            return val[:10] if val and val != 'nan' else ''
        elif 'tiktok' in nick or 'lazada' in nick:
            return add_one_day(row['ordered_date'])
        return ''

    filtered['MP SLA'] = filtered.apply(get_mp_sla, axis=1)
    filtered['WMS']    = filtered['invoice_number'].map(wms_map).fillna('NOT FOUND')
    filtered['_stock'] = filtered['custom_sku'].apply(lambda x: get_stock(x, inv_map))

    partial_none_mask  = filtered['WMS'].isin(['PARTIAL_ALLOCATED', 'NONE_ALLOCATED'])
    zero_stock_partial = partial_none_mask & (filtered['_stock'] == 0)

    cols_map = {
        'order_number':    'Order Number',
        'invoice_number':  'Invoice Number',
        'payment_status':  'Payment Status',
        'order_item_status': 'Order Item Status',
        'ordered_date':    'Ordered Date',
        'nickname':        'Nickname',
        'MP SLA':          'MP SLA',
        'WMS':             'WMS',
    }
    result = filtered[list(cols_map.keys())].rename(columns=cols_map).copy()
    result['_stock']      = filtered['_stock'].values
    result['_custom_sku'] = filtered['custom_sku'].values

    partial_none_mask_r  = result['WMS'].isin(['PARTIAL_ALLOCATED', 'NONE_ALLOCATED'])
    zero_stock_partial_r = partial_none_mask_r & (result['_stock'] == 0)

    not_pushed     = result[result['WMS'] == 'NOT FOUND'].copy()
    partial_all    = result[partial_none_mask_r].copy()
    close_nf       = result[result['WMS'].isin(['CLOSE', 'NOT FOUND'])].copy()
    # CANCEL excluded from FILTERED DATA + PIVOT
    filtered_data  = result[
        (result['WMS'] != 'NOT FOUND') &
        (result['WMS'] != 'CANCEL') &
        ~zero_stock_partial_r
    ].copy()

    rma_raw = tc[tc['order_status'].isin(['CANCELLED', 'CANCEL REQUESTED'])].copy()
    rma_raw['WMS'] = rma_raw['invoice_number'].map(wms_map).fillna('NOT FOUND')
    rma = rma_raw[rma_raw['WMS'] == 'FULFILLED'].drop_duplicates(subset=['invoice_number'])

    # ─ Build workbook ─
    today_str    = datetime.today().strftime('%Y-%m-%d')
    display_cols = ['Order Number', 'Invoice Number', 'Payment Status',
                    'Order Item Status', 'Ordered Date', 'Nickname', 'MP SLA', 'WMS']

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: FILTERED DATA
    ws1 = wb.create_sheet('FILTERED DATA')
    ws1.sheet_properties.tabColor = '4472C4'
    write_header(ws1, display_cols, [20, 20, 16, 20, 22, 30, 14, 22])
    final_rows = []
    for i, (_, row) in enumerate(filtered_data.iterrows(), 1):
        vals = [row['Order Number'], row['Invoice Number'], row['Payment Status'],
                row['Order Item Status'], row['Ordered Date'], row['Nickname'],
                row['MP SLA'], row['WMS']]
        final_rows.append(vals)
        apply_row_ws(ws1, i + 1, vals, get_row_fill(row['Nickname'], i), 8, row['WMS'])

    # Sheet 2: PIVOT
    ws2 = wb.create_sheet('PIVOT')
    ws2.sheet_properties.tabColor = '4472C4'
    headers_list = display_cols
    build_pivot_sheet(ws2, final_rows, headers_list, today_str)

    # Sheet 3: NOT PUSHED
    ws3 = wb.create_sheet('NOT PUSHED')
    ws3.sheet_properties.tabColor = 'FF0000'
    write_header(ws3, display_cols, [20, 20, 16, 20, 22, 30, 14, 22])
    for i, (_, row) in enumerate(not_pushed.iterrows(), 1):
        vals = [row['Order Number'], row['Invoice Number'], row['Payment Status'],
                row['Order Item Status'], row['Ordered Date'], row['Nickname'],
                row['MP SLA'], row['WMS']]
        apply_row_ws(ws3, i + 1, vals, get_row_fill(row['Nickname'], i), 8, row['WMS'])

    # Sheet 4: PARTIAL & NON ALLOCATED
    ws4 = wb.create_sheet('PARTIAL & NON ALLOCATED')
    ws4.sheet_properties.tabColor = 'ED7D31'
    p_cols = display_cols + ['Custom SKU', 'WH Stock', 'Remarks']
    write_header(ws4, p_cols, [20, 20, 16, 20, 22, 30, 14, 22, 22, 12, 12])
    for i, (_, row) in enumerate(partial_all.iterrows(), 1):
        raw_sku    = row['_custom_sku']
        sku_base   = raw_sku.split('x')[0] if raw_sku else ''
        custom_sku = f'GSK_{sku_base}' if sku_base else ''
        stock      = row['_stock']
        remarks    = 'CANCEL' if stock == 0 else ('FULFILL' if stock > 0 else '')
        rm_fill    = RED_FILL if remarks == 'CANCEL' else (GREEN_FILL if remarks == 'FULFILL' else PatternFill())
        vals       = [row['Order Number'], row['Invoice Number'], row['Payment Status'],
                      row['Order Item Status'], row['Ordered Date'], row['Nickname'],
                      row['MP SLA'], row['WMS'], custom_sku, stock if stock >= 0 else '', remarks]
        rf = get_row_fill(row['Nickname'], i)
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(i + 1, ci, v)
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
            if ci == 8 and row['WMS'] in WMS_FILLS:
                c.fill = WMS_FILLS[row['WMS']]
            elif ci == 11:
                c.fill = rm_fill
            else:
                c.fill = rf

    # Sheet 5: CLOSE & NOT FOUND
    ws5 = wb.create_sheet('CLOSE & NOT FOUND')
    ws5.sheet_properties.tabColor = '808080'
    write_header(ws5, display_cols, [20, 20, 16, 20, 22, 30, 14, 22])
    for i, (_, row) in enumerate(close_nf.iterrows(), 1):
        vals = [row['Order Number'], row['Invoice Number'], row['Payment Status'],
                row['Order Item Status'], row['Ordered Date'], row['Nickname'],
                row['MP SLA'], row['WMS']]
        apply_row_ws(ws5, i + 1, vals, get_row_fill(row['Nickname'], i), 8, row['WMS'])

    # Sheet 6: ORIGINAL
    ws6 = wb.create_sheet('ORIGINAL (All Data)')
    ws6.sheet_properties.tabColor = '1F1F1F'
    all_cols = list(tc.columns)
    write_header(ws6, all_cols)
    for ri, (_, row) in enumerate(tc.iterrows(), 2):
        for ci, col in enumerate(all_cols, 1):
            c = ws6.cell(ri, ci, row[col])
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER

    # Sheet 7: RMA
    ws7 = wb.create_sheet('RMA')
    ws7.sheet_properties.tabColor = 'C00000'
    rma_hdr = ['Order Number', 'Invoice Number', 'Order Status', 'Payment Status',
                'Ordered Date', 'Nickname', 'WMS']
    write_header(ws7, rma_hdr, [20, 20, 20, 16, 22, 30, 22])
    for i, (_, row) in enumerate(rma.iterrows(), 1):
        vals = [row['order_number'], row['invoice_number'], row['order_status'],
                row['payment_status'], row['ordered_date'], row['nickname'], row['WMS']]
        apply_row_ws(ws7, i + 1, vals, get_row_fill(row['nickname'], i), 7, row['WMS'])

    summary = {
        'FILTERED DATA':           len(filtered_data),
        'NOT PUSHED':              len(not_pushed),
        'PARTIAL & NON ALLOCATED': len(partial_all),
        'CLOSE & NOT FOUND':       len(close_nf),
        'ORIGINAL':                len(tc),
        'RMA':                     len(rma),
    }
    return wb, summary


# ─── NOT PUSHED UPDATER ──────────────────────────────────────────────────────

def update_not_pushed(wb_bytes, invoice_inputs, wms_status='PENDING_FULFILLMENT'):
    """
    Takes existing report bytes + list of invoice numbers.
    Moves matching rows from NOT PUSHED (and CLOSE & NOT FOUND) → FILTERED DATA,
    sets WMS = wms_status, rebuilds PIVOT. Returns updated wb bytes.
    """
    wb        = load_workbook(BytesIO(wb_bytes))
    today_str = datetime.today().strftime('%Y-%m-%d')

    invoices = {str(i).strip().upper() for i in invoice_inputs if str(i).strip()}

    DISPLAY_COLS = ['Order Number', 'Invoice Number', 'Payment Status',
                    'Order Item Status', 'Ordered Date', 'Nickname', 'MP SLA', 'WMS']
    WMS_COL_IDX  = 8  # 1-based

    def read_sheet_rows(ws):
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            rv = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in rv):
                rows.append(rv)
        return headers, rows

    def write_sheet_rows(ws, rows, headers):
        """Clear data rows and rewrite."""
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell       = ws.cell(r, c)
                cell.value = None
                cell.fill  = PatternFill()
                cell.font  = BODY_FONT
                cell.border = BORDER
        for i, rv in enumerate(rows, 1):
            nick    = rv[headers.index('Nickname')] if 'Nickname' in headers else ''
            wms_val = rv[headers.index('WMS')]      if 'WMS' in headers else None
            rf      = get_row_fill(nick, i)
            apply_row_ws(ws, i + 1, rv, rf, WMS_COL_IDX, wms_val)

    # 1. Read FILTERED DATA
    ws_fd       = wb['FILTERED DATA']
    fd_headers, fd_rows = read_sheet_rows(ws_fd)

    # 2. Read NOT PUSHED — separate matched vs kept
    ws_np       = wb['NOT PUSHED']
    np_headers, np_rows = read_sheet_rows(ws_np)
    inv_col_np  = np_headers.index('Invoice Number')

    matched_rows, kept_np_rows = [], []
    for rv in np_rows:
        inv_val = str(rv[inv_col_np] or '').strip().upper()
        if inv_val in invoices:
            updated    = list(rv[:WMS_COL_IDX - 1]) + [wms_status]
            matched_rows.append(updated)
        else:
            kept_np_rows.append(rv)

    # 3. Read CLOSE & NOT FOUND — remove matched
    ws_cnf        = wb['CLOSE & NOT FOUND']
    cnf_headers, cnf_rows = read_sheet_rows(ws_cnf)
    inv_col_cnf   = cnf_headers.index('Invoice Number')
    kept_cnf_rows = [rv for rv in cnf_rows
                     if str(rv[inv_col_cnf] or '').strip().upper() not in invoices]

    # 4. Append matched rows to FILTERED DATA
    final_fd_rows = fd_rows + matched_rows

    # 5. Rewrite all sheets
    write_sheet_rows(ws_fd,  final_fd_rows, fd_headers)
    write_sheet_rows(ws_np,  kept_np_rows,  np_headers)
    write_sheet_rows(ws_cnf, kept_cnf_rows, cnf_headers)

    # 6. Rebuild PIVOT
    ws_piv = wb['PIVOT']
    build_pivot_sheet(ws_piv, final_fd_rows, fd_headers, today_str)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), len(matched_rows), invoices - {
        str(r[inv_col_np]).strip().upper() for r in matched_rows
    }


# ─── STREAMLIT UI ────────────────────────────────────────────────────────────

st.title("📦 GSK Pending Order Report")
st.markdown("---")

tab1, tab2 = st.tabs(["🗂️ Generate Report", "✏️ Update NOT PUSHED Orders"])

# ══════════════════════════════════════════
#  TAB 1 — GENERATE REPORT
# ══════════════════════════════════════════
with tab1:
    st.subheader("Upload daily files to generate the report")

    col1, col2 = st.columns(2)
    with col1:
        tc_file     = st.file_uploader("TC Order Report (.csv)",        type=['csv'],  key='tc')
        shopee_file = st.file_uploader("Shopee Order File (.xlsx)",     type=['xlsx'], key='shopee')
    with col2:
        wms_files   = st.file_uploader("WMS File(s) (.xlsx)",           type=['xlsx'], key='wms',   accept_multiple_files=True)
        inv_file    = st.file_uploader("Inventory Report (.xlsx)",      type=['xlsx'], key='inv')

    if st.button("🚀 Generate Report", type='primary', use_container_width=True):
        if not all([tc_file, shopee_file, wms_files, inv_file]):
            st.error("Please upload all required files before generating the report.")
        else:
            with st.spinner("Processing…"):
                try:
                    wb, summary = generate_report(tc_file, shopee_file, wms_files, inv_file)

                    # Save to bytes
                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    report_bytes = buf.read()

                    # Store in session for updater tab
                    st.session_state['report_bytes']   = report_bytes
                    st.session_state['report_filename'] = f"GSK_TC_ORDER_REPORT_PROCESSED_{datetime.today().strftime('%Y%m%d')}.xlsx"

                    st.success("✅ Report generated successfully!")

                    # Summary table
                    st.markdown("#### Summary")
                    cols = st.columns(len(summary))
                    colors = {
                        'FILTERED DATA': '🔵', 'NOT PUSHED': '🔴',
                        'PARTIAL & NON ALLOCATED': '🟠', 'CLOSE & NOT FOUND': '⚫',
                        'ORIGINAL': '⬛', 'RMA': '🔴'
                    }
                    for col, (k, v) in zip(cols, summary.items()):
                        col.metric(f"{colors.get(k,'')} {k}", v)

                    st.download_button(
                        label     = "⬇️ Download Report",
                        data      = report_bytes,
                        file_name = st.session_state['report_filename'],
                        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                    if summary['NOT PUSHED'] > 0:
                        st.info(f"💡 **{summary['NOT PUSHED']} NOT PUSHED order(s)** — go to the **Update NOT PUSHED Orders** tab to update them once confirmed in WMS.")

                except Exception as e:
                    st.error(f"Error generating report: {e}")
                    st.exception(e)

# ══════════════════════════════════════════
#  TAB 2 — UPDATE NOT PUSHED ORDERS
# ══════════════════════════════════════════
with tab2:
    st.subheader("Update NOT PUSHED orders confirmed in WMS")
    st.markdown(
        "Upload the existing report (or use the one just generated), "
        "then type or paste the invoice numbers that are now in WMS."
    )

    # Load from session or file upload
    uploaded_report = st.file_uploader(
        "Upload existing report (.xlsx)  *(skip if you just generated it above)*",
        type=['xlsx'], key='existing_report'
    )

    if uploaded_report:
        st.session_state['report_bytes']   = uploaded_report.read()
        st.session_state['report_filename'] = uploaded_report.name

    has_report = 'report_bytes' in st.session_state and st.session_state['report_bytes']

    if has_report:
        st.success(f"📄 Report loaded: **{st.session_state.get('report_filename', 'report')}**")
    else:
        st.warning("No report loaded yet. Generate one in Tab 1 or upload an existing report above.")

    st.markdown("---")
    st.markdown("#### Enter invoice numbers (one per line, or comma-separated)")

    invoice_text = st.text_area(
        label       = "Invoice Numbers",
        placeholder = "e.g.\nGSK00352711\nGSK00352812\nGSK00353001",
        height      = 160,
        key         = 'invoice_input',
    )

    wms_status = st.selectbox(
        "WMS Status to assign",
        options=['PENDING_FULFILLMENT', 'FULFILLED', 'PARTIAL_ALLOCATED'],
        index=0,
        key='wms_status_select',
    )

    if st.button("✅ Update Report", type='primary', use_container_width=True, disabled=not has_report):
        raw_text  = invoice_text.strip()
        if not raw_text:
            st.error("Please enter at least one invoice number.")
        else:
            # Parse: split by newline or comma
            parts    = re.split(r'[\n,]+', raw_text)
            invoices = [p.strip() for p in parts if p.strip()]

            with st.spinner(f"Updating {len(invoices)} invoice(s)…"):
                try:
                    updated_bytes, matched_count, not_found = update_not_pushed(
                        st.session_state['report_bytes'],
                        invoices,
                        wms_status,
                    )

                    # Save updated bytes back to session
                    st.session_state['report_bytes'] = updated_bytes

                    if matched_count > 0:
                        st.success(
                            f"✅ **{matched_count} invoice(s)** moved from NOT PUSHED → FILTERED DATA "
                            f"with status `{wms_status}`. PIVOT updated."
                        )
                    else:
                        st.warning("No matching invoices found in NOT PUSHED sheet.")

                    if not_found:
                        st.warning(f"⚠️ These invoice(s) were not found in NOT PUSHED: {', '.join(sorted(not_found))}")

                    st.download_button(
                        label     = "⬇️ Download Updated Report",
                        data      = updated_bytes,
                        file_name = st.session_state.get('report_filename', 'GSK_REPORT_UPDATED.xlsx'),
                        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"Error updating report: {e}")
                    st.exception(e)
